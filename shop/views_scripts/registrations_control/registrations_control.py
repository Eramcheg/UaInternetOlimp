from io import BytesIO
from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook

import firebase_admin
from firebase_admin import firestore
from shop.views import school_registrations_ref
OBLAST_CHOICES = {"24": "Вінницька", "07": "Волинська", "12": "Дніпропетровська",
    "14": "Донецька", "18": "Житомирська", "21": "Закарпатська",
    "23": "Запорізька", "26": "Івано-Франківська", "32": "Київська",
    "35": "Кіровоградська", "44": "Луганська", "46": "Львівська",
    "48": "Миколаївська", "51": "Одеська", "53": "Полтавська",
    "56": "Рівненська", "59": "Сумська", "61": "Тернопільська",
    "63": "Харківська", "65": "Херсонська", "68": "Хмельницька",
    "71": "Черкаська", "73": "Чернівецька", "74": "Чернігівська",
    "80": "м. Київ"
}
COLUMNS = [
    ("firstName_uk",  "Ім'я (укр)"),
    ("lastName_uk",   "Прізвище (укр)"),
    ("patronymic_uk", "По батькові (укр)"),
    ("firstName_en",  "Ім'я (EN)"),
    ("lastName_en",   "Прізвище (EN)"),

    ("studyInUkraine", "Навчається в Україні"),
    ("schoolOblast", "Область школи"),
    ("schoolNumber", "Номер школи"),
    ("schoolName", "Назва школи"),

    ("residenceCountry", "Країна проживання"),
    ("residenceCity", "Місто проживання"),

    ("contactEmail", "Контактний емейл"),
    ("contactLinks", "Контактні посилання"),

    ("paralel", "Бажана паралель"),

    ("olympiadsParticipation", "Участь в олімпіадах"),
    ("olympiadsAchievements", "Досягнення в олімпіадах"),
]


def bool_to_text(val):
    if val is True:
        return "Так"
    elif val is False:
        return "Ні"
    return ""

def oblast_name(val):
    if val in OBLAST_CHOICES:
        return OBLAST_CHOICES[val]
    return val or ""
FIELD_TRANSFORMERS = {
    "studyInUkraine": bool_to_text,
    "olympiadsParticipation": bool_to_text,
    "schoolOblast": oblast_name,
}
def download_registrations(request):
    """
    GET /at/registrations_control/download_xlsx/
    Возвращает xlsx-файл с данными из коллекции 'school_registrations'.
    """
    # 1) Читаем данные из Firestore
    docs = school_registrations_ref.stream()

    # 2) Собираем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    # Заголовок
    ws.append([header for _, header in COLUMNS])

    for snap in docs:
        data = snap.to_dict() or {}
        row = []
        for field, _ in COLUMNS:
            val = data.get(field, "")
            # применяем трансформацию, если есть
            if field in FIELD_TRANSFORMERS:
                val = FIELD_TRANSFORMERS[field](val)
            row.append(val)
        ws.append(row)

    try:
        from openpyxl.utils import get_column_letter
        for col_idx, (_field, header) in enumerate(COLUMNS, start=1):
            max_len = len(str(header))
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
                for v in cell:
                    max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    except Exception:
        # Если что-то пойдёт не так — просто пропустим автоширину
        pass

    # 3) Отдаём как файл
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
